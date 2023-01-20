import openpyxl as xl

""" Here is where we declare variables and instantiate our arrays"""
wb = xl.load_workbook(filename='dndloot.xlsx', data_only=True)
ws = wb['Loot']
row_range = ws['M42'].value
gemArray = []
d12Array = []

""" This is finding the array that will house the list of gems. 
    These numbers should coordinate with the numbers rolled on the d12 """
for col in ws.iter_cols(min_row=4, min_col=14, max_row=(3 + row_range), max_col=14):
    for cell in col:
        gemArray.append(cell.value)
        # print(cell.value) """ This was a test """

""" This is were get the d12 Array. This is the rolled numbers off of a d12 for loot """
for col in ws.iter_cols(min_row=4, min_col=15, max_row=(4 + 11), max_col=15):
    for cell in col:
        d12Array.append(cell.value)

""" Here we display and check our work """
print(f"This is the gem array {gemArray}")
print(f"This is the d12 array {d12Array}")
print(f"There are {row_range} gems in this haul")

""" Here is where we check for duplicates """
gemList = []
for idx, i in enumerate(gemArray):
    if i in gemArray:
        gemCount = gemArray.count(i)
        gemList.append(f"{gemCount}:{i}")
        print(f"there are {gemCount} of {i}'s in this list")

finalGemList = []
[finalGemList.append(x) for x in gemList if x not in finalGemList]
lootList = []
print(finalGemList)
"""Please make a note the format finalGemList [##:##] 
    equates to [number of gems of that type:the gem number from the table."""


def getgemvalues(finalgemlist, listindex, numberofgems=None, numberofgemtype=None):
    sub_string = finalgemlist[listindex]
    findmiddle = sub_string.find(":")

    """If findmiddle is equal to 1, then we know that the 
        number of gems is a single digit."""
    if findmiddle == 1:
        numberofgems = sub_string[0]

        if len(sub_string) == 3:
            numberofgemtype = sub_string[-1]

        if len(sub_string) == 4:
            numberofgemtype = sub_string[2] + sub_string[3]

    if findmiddle != 1:
        numberofgems = sub_string[0] + sub_string[1]

        if len(sub_string) == 3:
            numberofgemtype = sub_string[-1]

        if len(sub_string) == 4:
            numberofgemtype = sub_string[3]

        if len(sub_string) == 5:
            numberofgemtype = sub_string[3] + sub_string[4]

    return numberofgems, numberofgemtype


"""Now lets grab our data from the spreadsheet"""


def readgems(rowrange=None, numberofgemtype=0, gemnumber=0):
    print("This is the gem number " + str(gemnumber))
    print("This is the number of gems of that type " + str(numberofgemtype))
    print("\n")

    if numberofgemtype > 1:
        for col in ws.iter_cols(min_row=3+gemnumber, min_col=16, max_row=3+gemnumber, max_col=16):
            for cell in col:
                counter = 0
                while counter < numberofgemtype:
                    lootList.append(cell.value)
                    counter += 1

    if numberofgemtype == 1:
        for col in ws.iter_cols(min_row=3+gemnumber, min_col=16, max_row=3+gemnumber, max_col=16):
            for cell in col:
                lootList.append(cell.value)

    return  lootList


workingList = getgemvalues(finalGemList, -1)

numberOfGems = workingList[0]
numberOfGemType = workingList[1]
numberOfGemType = int(numberOfGemType)
numberOfGems = int(numberOfGems)

print(numberOfGemType)
print(numberOfGems)

readgems(row_range, numberOfGems, numberOfGemType)
print(lootList)

lootList = str(lootList)

result = ws['h42']
result.value = lootList

#wb.save("C:\\Users\\samue\\PycharmProjects\\dndlootgenerator\\dndloot.xlsx")
print("saved")
